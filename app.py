import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import base64

# Page configuration
st.set_page_config(
    page_title="Ledger Management System",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'ledger_data' not in st.session_state:
    st.session_state.ledger_data = pd.DataFrame()
if 'initial_balance' not in st.session_state:
    st.session_state.initial_balance = 304205.0

def calculate_balance_and_type(df, initial_balance):
    """Calculate balance and type for each transaction"""
    df = df.copy()
    
    # Ensure Dr Amount and Cr Amount are numeric
    df['Dr Amount (₹)'] = pd.to_numeric(df['Dr Amount (₹)'], errors='coerce').fillna(0)
    df['Cr Amount (₹)'] = pd.to_numeric(df['Cr Amount (₹)'], errors='coerce').fillna(0)
    
    # Initialize balance and type columns
    df['Balance (₹)'] = 0.0
    df['Type (Dr/Cr)'] = ''
    
    current_balance = initial_balance
    
    for idx in range(len(df)):
        dr_amount = df.loc[idx, 'Dr Amount (₹)']
        cr_amount = df.loc[idx, 'Cr Amount (₹)']
        
        if dr_amount > 0:
            current_balance -= dr_amount
            df.loc[idx, 'Type (Dr/Cr)'] = 'Dr'
        elif cr_amount > 0:
            current_balance += cr_amount
            df.loc[idx, 'Type (Dr/Cr)'] = 'Cr'
        else:
            df.loc[idx, 'Type (Dr/Cr)'] = ''
        
        df.loc[idx, 'Balance (₹)'] = current_balance
    
    return df

def format_currency(value):
    """Format currency with Indian rupee symbol"""
    if pd.isna(value) or value == 0:
        return ""
    return f"₹{value:,.2f}"

def create_excel_download(df, initial_balance):
    """Create Excel file with proper formatting"""
    output = io.BytesIO()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    
    # Headers
    headers = ['Date', 'Particulars', 'C/F', 'Dr Amount (₹)', 'Cr Amount (₹)', 'Balance (₹)', 'Type (Dr/Cr)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Add initial balance row
    ws.cell(row=2, column=1, value="Opening Balance")
    ws.cell(row=2, column=2, value="Opening Balance")
    ws.cell(row=2, column=6, value=initial_balance)
    ws.cell(row=2, column=7, value="Opening")
    
    # Add data rows
    for idx, row in df.iterrows():
        excel_row = idx + 3  # +3 because of header and opening balance
        
        # Date
        if pd.notna(row['Date']):
            ws.cell(row=excel_row, column=1, value=row['Date'])
        
        # Particulars
        ws.cell(row=excel_row, column=2, value=row['Particulars'])
        
        # C/F (if exists)
        if 'C/F' in row and pd.notna(row['C/F']):
            ws.cell(row=excel_row, column=3, value=row['C/F'])
        
        # Dr Amount
        if row['Dr Amount (₹)'] > 0:
            ws.cell(row=excel_row, column=4, value=row['Dr Amount (₹)'])
        
        # Cr Amount
        if row['Cr Amount (₹)'] > 0:
            ws.cell(row=excel_row, column=5, value=row['Cr Amount (₹)'])
        
        # Balance
        ws.cell(row=excel_row, column=6, value=row['Balance (₹)'])
        
        # Type
        ws.cell(row=excel_row, column=7, value=row['Type (Dr/Cr)'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def main():
    st.title("📊 Ledger Management System")
    st.markdown("---")
    
    # Sidebar for settings
    st.sidebar.header("Settings")
    
    # Initial balance setting
    new_initial_balance = st.sidebar.number_input(
        "Initial Balance (₹)",
        value=st.session_state.initial_balance,
        step=1000.0,
        format="%.2f"
    )
    
    if new_initial_balance != st.session_state.initial_balance:
        st.session_state.initial_balance = new_initial_balance
        if not st.session_state.ledger_data.empty:
            st.session_state.ledger_data = calculate_balance_and_type(
                st.session_state.ledger_data, st.session_state.initial_balance
            )
    
    # Main content area
    tab1, tab2, tab3 = st.tabs(["📁 Upload Data", "✏️ Manual Entry", "📊 View Ledger"])
    
    with tab1:
        st.header("Upload Excel/CSV File")
        st.markdown("Upload your transaction data with columns: Date, Particulars, Dr Amount, Cr Amount")
        
        uploaded_file = st.file_uploader(
            "Choose a file",
            type=['xlsx', 'xls', 'csv'],
            help="Upload Excel or CSV file with transaction data"
        )
        
        if uploaded_file is not None:
            try:
                # Read file
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
                
                st.success(f"File uploaded successfully! Found {len(df)} rows.")
                
                # Display preview
                st.subheader("Data Preview")
                st.dataframe(df.head())
                
                # Column mapping
                st.subheader("Column Mapping")
                col1, col2 = st.columns(2)
                
                with col1:
                    date_col = st.selectbox("Date Column", df.columns.tolist(), index=0)
                    particulars_col = st.selectbox("Particulars Column", df.columns.tolist(), index=1)
                
                with col2:
                    dr_col = st.selectbox("Dr Amount Column", df.columns.tolist(), 
                                        index=2 if len(df.columns) > 2 else 0)
                    cr_col = st.selectbox("Cr Amount Column", df.columns.tolist(), 
                                        index=3 if len(df.columns) > 3 else 0)
                
                # Optional C/F column
                cf_col = st.selectbox("C/F Column (Optional)", ["None"] + df.columns.tolist())
                
                if st.button("Process Data", type="primary"):
                    # Create standardized dataframe
                    processed_df = pd.DataFrame()
                    processed_df['Date'] = pd.to_datetime(df[date_col], errors='coerce')
                    processed_df['Particulars'] = df[particulars_col].astype(str)
                    
                    if cf_col != "None":
                        processed_df['C/F'] = df[cf_col]
                    else:
                        processed_df['C/F'] = ''
                    
                    processed_df['Dr Amount (₹)'] = pd.to_numeric(df[dr_col], errors='coerce').fillna(0)
                    processed_df['Cr Amount (₹)'] = pd.to_numeric(df[cr_col], errors='coerce').fillna(0)
                    
                    # Calculate balance and type
                    processed_df = calculate_balance_and_type(processed_df, st.session_state.initial_balance)
                    
                    # Store in session state
                    st.session_state.ledger_data = processed_df
                    
                    st.success("Data processed successfully!")
                    st.rerun()
                    
            except Exception as e:
                st.error(f"Error processing file: {str(e)}")
    
    with tab2:
        st.header("Manual Transaction Entry")
        
        with st.form("manual_entry_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                entry_date = st.date_input("Date", value=datetime.now().date())
                particulars = st.text_input("Particulars", placeholder="Description of transaction")
                cf_entry = st.text_input("C/F (Optional)", placeholder="Carried Forward reference")
            
            with col2:
                transaction_type = st.radio("Transaction Type", ["Debit", "Credit"])
                amount = st.number_input("Amount (₹)", min_value=0.0, step=0.01, format="%.2f")
            
            submitted = st.form_submit_button("Add Transaction", type="primary")
            
            if submitted:
                if particulars and amount > 0:
                    # Create new transaction
                    new_transaction = {
                        'Date': pd.to_datetime(entry_date),
                        'Particulars': particulars,
                        'C/F': cf_entry,
                        'Dr Amount (₹)': amount if transaction_type == "Debit" else 0,
                        'Cr Amount (₹)': amount if transaction_type == "Credit" else 0
                    }
                    
                    # Add to existing data or create new
                    if st.session_state.ledger_data.empty:
                        st.session_state.ledger_data = pd.DataFrame([new_transaction])
                    else:
                        st.session_state.ledger_data = pd.concat([
                            st.session_state.ledger_data,
                            pd.DataFrame([new_transaction])
                        ], ignore_index=True)
                    
                    # Recalculate balance and type
                    st.session_state.ledger_data = calculate_balance_and_type(
                        st.session_state.ledger_data, st.session_state.initial_balance
                    )
                    
                    st.success("Transaction added successfully!")
                    st.rerun()
                else:
                    st.error("Please fill in all required fields and enter a valid amount.")
    
    with tab3:
        st.header("Ledger View")
        
        if not st.session_state.ledger_data.empty:
            # Display current balance
            current_balance = st.session_state.ledger_data['Balance (₹)'].iloc[-1]
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Initial Balance", f"₹{st.session_state.initial_balance:,.2f}")
            with col2:
                st.metric("Current Balance", f"₹{current_balance:,.2f}")
            with col3:
                net_change = current_balance - st.session_state.initial_balance
                st.metric("Net Change", f"₹{net_change:,.2f}", 
                         delta=f"₹{net_change:,.2f}")
            
            st.markdown("---")
            
            # Display data with formatting
            display_df = st.session_state.ledger_data.copy()
            display_df['Date'] = display_df['Date'].dt.strftime('%Y-%m-%d')
            
            # Format currency columns
            for col in ['Dr Amount (₹)', 'Cr Amount (₹)', 'Balance (₹)']:
                display_df[col] = display_df[col].apply(lambda x: f"₹{x:,.2f}" if x != 0 else "")
            
            # Display table
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Date": st.column_config.TextColumn("Date", width="medium"),
                    "Particulars": st.column_config.TextColumn("Particulars", width="large"),
                    "C/F": st.column_config.TextColumn("C/F", width="small"),
                    "Dr Amount (₹)": st.column_config.TextColumn("Dr Amount (₹)", width="medium"),
                    "Cr Amount (₹)": st.column_config.TextColumn("Cr Amount (₹)", width="medium"),
                    "Balance (₹)": st.column_config.TextColumn("Balance (₹)", width="medium"),
                    "Type (Dr/Cr)": st.column_config.TextColumn("Type", width="small")
                }
            )
            
            # Download button
            st.markdown("---")
            excel_data = create_excel_download(st.session_state.ledger_data, st.session_state.initial_balance)
            
            st.download_button(
                label="📥 Download Ledger as Excel",
                data=excel_data,
                file_name=f"ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
            
            # Clear data button
            if st.button("🗑️ Clear All Data", type="secondary"):
                st.session_state.ledger_data = pd.DataFrame()
                st.success("All data cleared!")
                st.rerun()
                
        else:
            st.info("No data available. Please upload a file or add transactions manually.")
            
            # Show sample format
            st.subheader("Sample Data Format")
            sample_data = pd.DataFrame({
                'Date': ['2024-01-01', '2024-01-02', '2024-01-03'],
                'Particulars': ['Opening Balance', 'Sales Revenue', 'Office Rent'],
                'Dr Amount (₹)': [0, 0, 25000],
                'Cr Amount (₹)': [0, 50000, 0]
            })
            st.dataframe(sample_data, use_container_width=True)

if __name__ == "__main__":
    main()